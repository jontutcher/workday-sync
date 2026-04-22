"""Parse a Workday 'My Absence' XLSX export into AbsenceRequest objects."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import NamedTuple

import openpyxl

from workday_sync.models import AbsenceRequest, HalfDayPeriod

# Row indices within the sheet (1-based, as openpyxl uses)
_TITLE_ROW = 1
_HEADER_ROW = 5
_DATA_START_ROW = 6

# Expected column headers (lowercase) mapped to field names
_COLUMN_MAP = {
    "date": "date",
    "type": "leave_type",
    "requested": "hours",
    "unit of time": "unit_of_time",
    "comment": "comment",
    "status": "status",
}

# Hours per working day — used to normalise Day-unit rows
_HOURS_PER_DAY = 8.0


class ParseResult(NamedTuple):
    """Return value of :func:`parse_xlsx`."""

    requests: list[AbsenceRequest]
    warnings: list[str]


def parse_xlsx(path: Path) -> ParseResult:
    """Parse a Workday absence XLSX export and return a ParseResult.

    Args:
        path: Path to the .xlsx file.

    Returns:
        A ParseResult containing the parsed requests and any user-facing
        warning messages (e.g. ambiguous half-day periods).

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file does not look like a Workday absence export,
            or contains an unsupported unit of time.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    user_name = _extract_user_name(ws)
    col_indices = _extract_column_indices(ws)

    requests: list[AbsenceRequest] = []
    parse_warnings: list[str] = []

    for row in ws.iter_rows(min_row=_DATA_START_ROW, values_only=True):
        if not any(row):
            break

        raw_date = row[col_indices["date"]]
        if raw_date is None:
            continue

        parsed_date = _parse_date(raw_date)
        raw_hours = float(row[col_indices["hours"]] or 0)
        unit = str(row[col_indices["unit_of_time"]] or "").strip()
        hours = _normalize_to_hours(raw_hours, unit)
        leave_type = str(row[col_indices["leave_type"]] or "")
        comment_raw = row[col_indices["comment"]]
        comment = str(comment_raw).strip() if comment_raw is not None else None
        status = str(row[col_indices["status"]] or "")

        req = AbsenceRequest(
            date=parsed_date,
            hours=hours,
            leave_type=leave_type,
            comment=comment,
            status=status,
            user_name=user_name,
        )

        if req.is_half_day and HalfDayPeriod.from_comment(comment) is HalfDayPeriod.UNKNOWN:
            parse_warnings.append(
                f"Half-day on {parsed_date} has no AM/PM indicator in comment "
                f"({comment!r}). Defaulting to morning (08:00–12:00)."
            )

        requests.append(req)

    return ParseResult(requests=requests, warnings=parse_warnings)


def _normalize_to_hours(value: float, unit: str) -> float:
    """Convert a Workday 'Requested' value to hours based on its unit.

    Args:
        value: The numeric quantity from the 'Requested' column.
        unit: The string from the 'Unit of Time' column (e.g. 'Hours', 'Days').

    Returns:
        The equivalent duration in hours.

    Raises:
        ValueError: If *unit* is not 'Hours' or 'Days'.
    """
    unit_lower = unit.lower()
    if unit_lower == "hours":
        return value
    if unit_lower == "days":
        return value * _HOURS_PER_DAY
    raise ValueError(
        f"Unsupported unit of time: {unit!r}. Expected 'Hours' or 'Days'."
    )


def _extract_user_name(ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
    """Extract user name from the title cell, e.g. 'My Absence: Jon Tutcher'."""
    title_cell = ws.cell(row=_TITLE_ROW, column=1).value or ""
    title = str(title_cell)
    if ":" in title:
        return title.split(":", 1)[1].strip()
    raise ValueError(f"Could not extract user name from title row: {title!r}")


def _extract_column_indices(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    """Return a dict mapping field names to 0-based column indices."""
    header_row = [
        (cell.value or "").strip().lower()
        for cell in ws[_HEADER_ROW]
    ]
    indices: dict[str, int] = {}
    for header, field in _COLUMN_MAP.items():
        try:
            indices[field] = header_row.index(header)
        except ValueError as e:
            raise ValueError(
                f"Expected column {header!r} not found in headers: {header_row}"
            ) from e
    return indices


def _parse_date(value: datetime | date | str) -> date:
    """Normalise an openpyxl date cell value to a plain date object."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise ValueError(f"Unexpected date value: {value!r}")
